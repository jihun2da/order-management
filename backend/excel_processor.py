# -*- coding: utf-8 -*-
"""
엑셀 처리 엔진 v3.0 (완전 로컬 주문번호 생성)
핵심 최적화:
  - generate_order_no RPC 완전 제거: buyer_consignor_counters를 Python에서 직접 관리
  - 34,000행 파일 예상 처리시간: 30~60초
  - 모든 INSERT/UPDATE는 400개씩 배치 처리
"""
import io
import re
import sys
from datetime import datetime
from collections import defaultdict
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import COLOR_INDEX

EXACT_RGB_MAP = {
    "FFFFFF00": "입고", "FFFF00": "입고",
    "FF00FFFF": "미송", "00FFFF": "미송",
    "FFFF0000": "품절", "FF0000": "품절",
    "FFFFC000": "교환", "FFC000": "교환",
    "FFE6B8B7": "환불", "E6B8B7": "환불",
    "FFBFBFBF": "택배비", "BFBFBF": "택배비",
}
THEME_PATTERN_MAP = {
    (0, -0.249977): "택배비",
    (5, 0.599994):  "환불",
}
DEFAULT_OFFICE_THEME = {
    0:(0,0,0), 1:(255,255,255), 2:(238,236,225),
    3:(31,73,125), 4:(79,129,189), 5:(192,80,77),
}
STATUS_COLORS = {
    "입고":"FFFFFF00","미송":"FF00FFFF","품절":"FFFF0000",
    "교환":"FFFFC000","환불":"FFE6B8B7","택배비":"FFBFBFBF",
}
BATCH = 400


def _log(msg):
    print(msg, flush=True)
    sys.stdout.flush()


# ── 색상 유틸 ──────────────────────────
def _color_to_rgb(c) -> Optional[str]:
    if not c: return None
    if c.type == "rgb":
        return c.rgb[2:] if len(c.rgb)==8 else c.rgb
    if c.type == "theme":
        tid, tint = c.theme, round(c.tint or 0.0, 6)
        if (tid, tint) in THEME_PATTERN_MAP: return None
        if tid in DEFAULT_OFFICE_THEME:
            return "%02x%02x%02x" % DEFAULT_OFFICE_THEME[tid]
    if c.type == "indexed":
        try:
            rgb = COLOR_INDEX[c.indexed]
            return rgb[2:] if len(rgb)==8 else rgb
        except: pass
    return None

def _get_cell_status(cell) -> str:
    if not cell.fill or not cell.fill.start_color: return "입고대기"
    color = cell.fill.start_color
    if color.type == "theme":
        key = (color.theme, round(color.tint or 0.0, 6))
        if key in THEME_PATTERN_MAP: return THEME_PATTERN_MAP[key]
    rgb = _color_to_rgb(color)
    if rgb: return EXACT_RGB_MAP.get(rgb.upper(), "입고대기")
    return "입고대기"

def _parse_date(val):
    if isinstance(val, datetime): return val.date().isoformat()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d"):
            try: return datetime.strptime(val, fmt).date().isoformat()
            except: pass
    if isinstance(val, (int, float)):
        try: return datetime.strptime(str(int(val)), "%Y%m%d").date().isoformat()
        except: pass
    return datetime.today().date().isoformat()

def _extract_manager_code(val) -> str:
    if val and isinstance(val, str):
        m = re.match(r"^([A-Za-z]+)", val.strip())
        if m: return m.group(1).upper()[:2]
    return "XX"

def _get_col_map(ws) -> dict:
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v: col_map[str(v).strip()] = c
    def fc(n, d): return col_map.get(n, d)
    return {
        "manager": fc("알파벳",1), "barcode": fc("미등록주문",2),
        "order_date": fc("주문일",3), "user_id": 4,
        "order_no": fc("고유번호",5), "buyer": fc("주문자명",6),
        "consignor": fc("위탁자명",7), "brand": fc("브랜드",8),
        "product": fc("상품명",9), "color": fc("색상",10),
        "size": fc("사이즈",11), "quantity": fc("수량",12),
        "options": fc("상가",13), "wholesale": fc("도매가",14),
        "supplier": fc("미송",15), "notes": fc("비고",16),
        "recipient_name": fc("이름",17), "phone": fc("전화번호",18),
        "address": fc("주소",19), "buyer_user_id": 20,
        "delivery_msg": fc("배송메세지",21), "code": fc("코드",22),
    }

def _val(ws, row, col):
    v = ws.cell(row, col).value
    return "" if v is None else str(v).strip()

def _chunks(lst, n):
    for i in range(0, len(lst), n): yield lst[i:i+n]

def _bulk_insert(supabase, table, rows):
    for chunk in _chunks(rows, BATCH):
        supabase.table(table).insert(chunk).execute()

def _bulk_fetch_orders(supabase, nos) -> dict:
    result = {}
    for chunk in _chunks(nos, BATCH):
        data = supabase.table("orders").select("id,order_no,manager_id").in_(
            "order_no", chunk).execute().data or []
        for r in data: result[r["order_no"]] = r
    return result

def _bulk_fetch_items(supabase, ids) -> dict:
    result = defaultdict(list)
    for chunk in _chunks(ids, BATCH):
        data = supabase.table("order_items").select(
            "id,order_id,product_name,status,quantity,color,status_history,change_log"
        ).in_("order_id", chunk).execute().data or []
        for r in data: result[r["order_id"]].append(r)
    return result


# ── 엔티티 캐시 빌더 ──────────────────
def _build_entity_caches(supabase, raw_rows):
    # ─ managers ─
    mgr_codes = list({r["manager_code"] for r in raw_rows})
    existing_mgrs = supabase.table("managers").select("id,code").in_(
        "code", mgr_codes).execute().data or []
    mgr_cache = {m["code"]: m["id"] for m in existing_mgrs}
    missing_mgrs = [{"code": c, "name": c} for c in mgr_codes if c not in mgr_cache]
    if missing_mgrs:
        for chunk in _chunks(missing_mgrs, BATCH):
            res = supabase.table("managers").insert(chunk).execute()
            if res.data:
                for m in res.data: mgr_cache[m["code"]] = m["id"]
        if len(mgr_cache) < len(mgr_codes):
            refetch = supabase.table("managers").select("id,code").in_(
                "code", mgr_codes).execute().data or []
            for m in refetch: mgr_cache[m["code"]] = m["id"]

    # ─ consignors ─
    con_names = list({r["consignor_name"] for r in raw_rows if r["consignor_name"]})
    existing_cons = []
    for chunk in _chunks(con_names, BATCH):
        existing_cons += supabase.table("consignors").select("id,name").in_(
            "name", chunk).execute().data or []
    con_cache = {c["name"]: c["id"] for c in existing_cons}
    missing_cons = [{"name": n} for n in con_names if n not in con_cache]
    if missing_cons:
        for chunk in _chunks(missing_cons, BATCH):
            res = supabase.table("consignors").insert(chunk).execute()
            if res.data:
                for c in res.data: con_cache[c["name"]] = c["id"]
        if len(con_cache) < len(con_names):
            for chunk in _chunks(con_names, BATCH):
                rf = supabase.table("consignors").select("id,name").in_(
                    "name", chunk).execute().data or []
                for c in rf: con_cache[c["name"]] = c["id"]

    # ─ buyers ─
    buyer_uid_map = {}
    buyer_name_map = {}
    for r in raw_rows:
        uid = r["buyer_user_id"]
        if uid:
            if uid not in buyer_uid_map: buyer_uid_map[uid] = r
        else:
            name = r["buyer_name"]
            if name not in buyer_name_map: buyer_name_map[name] = r

    buyer_cache = {}

    if buyer_uid_map:
        uids = list(buyer_uid_map.keys())
        for chunk in _chunks(uids, BATCH):
            data = supabase.table("buyers").select("id,user_id,name").in_(
                "user_id", chunk).execute().data or []
            for b in data: buyer_cache[b["user_id"]] = b["id"]

    if buyer_name_map:
        names = list(buyer_name_map.keys())
        for chunk in _chunks(names, BATCH):
            data = supabase.table("buyers").select("id,name").in_(
                "name", chunk).execute().data or []
            for b in data: buyer_cache[b["name"]] = b["id"]

    missing_uid_buyers = [
        {"name": r["buyer_name"], "user_id": uid, "phone": r["phone"] or None}
        for uid, r in buyer_uid_map.items() if uid not in buyer_cache
    ]
    if missing_uid_buyers:
        for chunk in _chunks(missing_uid_buyers, BATCH):
            res = supabase.table("buyers").insert(chunk).execute()
            if res.data:
                for b in res.data:
                    if b.get("user_id"): buyer_cache[b["user_id"]] = b["id"]
        missing_uids = [uid for uid in buyer_uid_map if uid not in buyer_cache]
        if missing_uids:
            for chunk in _chunks(missing_uids, BATCH):
                rf = supabase.table("buyers").select("id,user_id").in_(
                    "user_id", chunk).execute().data or []
                for b in rf: buyer_cache[b["user_id"]] = b["id"]

    missing_name_buyers = [
        {"name": name, "user_id": None, "phone": r["phone"] or None}
        for name, r in buyer_name_map.items() if name not in buyer_cache
    ]
    if missing_name_buyers:
        for chunk in _chunks(missing_name_buyers, BATCH):
            res = supabase.table("buyers").insert(chunk).execute()
            if res.data:
                for b in res.data: buyer_cache[b["name"]] = b["id"]
        missing_names = [n for n in buyer_name_map if n not in buyer_cache]
        if missing_names:
            for chunk in _chunks(missing_names, BATCH):
                rf = supabase.table("buyers").select("id,name").in_(
                    "name", chunk).execute().data or []
                for b in rf: buyer_cache[b["name"]] = b["id"]

    return mgr_cache, buyer_cache, con_cache


def _get_buyer_id(row, buyer_cache):
    uid = row["buyer_user_id"]
    if uid and uid in buyer_cache: return buyer_cache[uid]
    name = row["buyer_name"]
    if name in buyer_cache: return buyer_cache[name]
    return None


# ── 로컬 주문번호 생성 (RPC 없이) ─────
def _build_order_no_engine(supabase, groups):
    """
    buyer_consignor_counters 테이블을 직접 읽어 기존 base_number를 로드하고,
    새 그룹에는 순차적으로 번호를 할당한 뒤 bulk INSERT 한다.
    반환: group_key → base_number
    Returns: {group_key: int}
    """
    # 그룹별 key 목록
    group_list = list(groups.items())  # [(group_key, {buyer_id, consignor_id, manager_code}), ...]

    # 1. 기존 카운터 전체 fetch (작은 테이블)
    existing = supabase.table("buyer_consignor_counters").select(
        "buyer_id,consignor_id,manager_code,base_number"
    ).execute().data or []

    counter_map = {}  # (buyer_id, consignor_id_or_None, manager_code) → base_number
    for row in existing:
        key = (str(row["buyer_id"]), str(row["consignor_id"]) if row["consignor_id"] else None, row["manager_code"])
        counter_map[key] = row["base_number"]

    # 2. 재활용 번호 fetch
    recycled = supabase.table("completed_order_numbers").select(
        "id,manager_code,base_number"
    ).order("completed_at", desc=False).execute().data or []

    recycled_by_mgr = defaultdict(list)
    for r in recycled:
        recycled_by_mgr[r["manager_code"]].append((r["id"], r["base_number"]))

    # 3. 매니저별 현재 최대 번호 추적
    max_by_mgr = defaultdict(int)
    for (_, _, mc), bn in counter_map.items():
        if bn > max_by_mgr[mc]: max_by_mgr[mc] = bn

    # 4. 각 그룹에 base_number 할당
    result = {}
    new_counters = []    # DB에 INSERT할 새 카운터
    recycled_to_delete = []  # 재활용 후 삭제할 ID

    for gk, gdata in group_list:
        buyer_id   = str(gdata["buyer_id"])
        con_id_raw = gdata["consignor_id"]
        con_id     = str(con_id_raw) if con_id_raw else None
        mc         = gdata["manager_code"]

        lookup_key = (buyer_id, con_id, mc)
        if lookup_key in counter_map:
            result[gk] = counter_map[lookup_key]
        else:
            # 재활용 풀 확인
            if recycled_by_mgr[mc]:
                rid, rbn = recycled_by_mgr[mc].pop(0)
                base_num = rbn
                recycled_to_delete.append(rid)
            else:
                max_by_mgr[mc] += 1
                base_num = max_by_mgr[mc]

            result[gk] = base_num
            counter_map[lookup_key] = base_num
            new_counters.append({
                "buyer_id":     buyer_id,
                "consignor_id": con_id,
                "manager_code": mc,
                "base_number":  base_num,
            })

    # 5. 새 카운터 bulk INSERT
    if new_counters:
        for chunk in _chunks(new_counters, BATCH):
            try:
                supabase.table("buyer_consignor_counters").insert(chunk).execute()
            except Exception as e:
                _log(f"[WARN] counter insert error: {e}")

    # 6. 재활용된 번호 삭제
    if recycled_to_delete:
        for chunk in _chunks(recycled_to_delete, BATCH):
            try:
                supabase.table("completed_order_numbers").delete().in_(
                    "id", chunk).execute()
            except Exception as e:
                _log(f"[WARN] recycled delete error: {e}")

    return result


# ──────────────────────────────────────
# 메인 처리 함수
# ──────────────────────────────────────
def process_excel_file(file_contents: bytes, filename: str, supabase,
                       pre_upload_id: str = None) -> dict:
    upload_id = pre_upload_id
    try:
        if not upload_id:
            hist = supabase.table("upload_history").insert({
                "filename": filename, "status": "처리중"
            }).execute()
            upload_id = hist.data[0]["id"]

        _log(f"[PROC] 시작: {filename}, upload_id={upload_id}")

        wb = load_workbook(io.BytesIO(file_contents))
        ws = wb.active
        col = _get_col_map(ws)

        # ─── 1단계: 행 파싱 ───
        _log(f"[PROC] 행 파싱 중... (총 {ws.max_row}행)")
        raw_rows = []
        for r in range(2, ws.max_row + 1):
            buyer_name = _val(ws, r, col["buyer"])
            prod_name  = _val(ws, r, col["product"])
            if not buyer_name or not prod_name: continue
            raw_rows.append({
                "row_idx":           r,
                "existing_order_no": _val(ws, r, col["order_no"]).strip(),
                "manager_code":      _extract_manager_code(_val(ws, r, col["manager"])),
                "buyer_name":        buyer_name,
                "buyer_user_id":     _val(ws, r, col["user_id"]),
                "consignor_name":    _val(ws, r, col["consignor"]),
                "order_date":        _parse_date(ws.cell(r, col["order_date"]).value),
                "product_name":      prod_name,
                "status":            _get_cell_status(ws.cell(r, col["product"])),
                "quantity":          int(ws.cell(r, col["quantity"]).value or 1),
                "color":             _val(ws, r, col["color"]),
                "barcode":           _val(ws, r, col["barcode"]),
                "brand":             _val(ws, r, col["brand"]),
                "size":              _val(ws, r, col["size"]),
                "options":           _val(ws, r, col["options"]),
                "wholesale":         _val(ws, r, col["wholesale"]),
                "supplier":          _val(ws, r, col["supplier"]),
                "item_notes":        _val(ws, r, col["notes"]),
                "recipient_name":    _val(ws, r, col["recipient_name"]),
                "phone":             _val(ws, r, col["phone"]),
                "address":           _val(ws, r, col["address"]),
                "delivery_msg":      _val(ws, r, col["delivery_msg"]),
                "item_code":         _val(ws, r, col["code"]),
                "bx_user_id":        _val(ws, r, col["buyer_user_id"]),
            })

        wb.close()  # 메모리 해제
        _log(f"[PROC] 파싱 완료: {len(raw_rows)}행")

        if not raw_rows:
            supabase.table("upload_history").update(
                {"status":"완료","rows_processed":0}).eq("id", upload_id).execute()
            return {"success":True,"upload_id":upload_id,"inserted":0,"updated":0,"errors":[]}

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        # ─── 2단계: 신규 행 그룹화 ───
        new_groups: dict = defaultdict(list)
        for row in raw_rows:
            if not row["existing_order_no"]:
                key = (row["buyer_name"],row["consignor_name"],row["order_date"],row["manager_code"])
                new_groups[key].append(row)
        for (bn,cn,od,mc), items in new_groups.items():
            total = len(items)
            for idx, item in enumerate(items, 1):
                item["seq_num"] = idx
                item["total_count"] = total
                item["is_consignment"] = (bn == cn)

        # ─── 3단계: 엔티티 캐시 빌드 ───
        _log("[PROC] 엔티티 캐시 빌드 중...")
        mgr_cache, buyer_cache, con_cache = _build_entity_caches(supabase, raw_rows)
        _log(f"[PROC] 엔티티 캐시 완료: mgr={len(mgr_cache)}, buyer={len(buyer_cache)}, con={len(con_cache)}")

        # ─── 4단계: 기존 주문 일괄 조회 ───
        all_nos = list({r["existing_order_no"] for r in raw_rows if r["existing_order_no"]})
        order_map = _bulk_fetch_orders(supabase, all_nos) if all_nos else {}
        order_ids = [o["id"] for o in order_map.values()]
        item_map  = _bulk_fetch_items(supabase, order_ids) if order_ids else {}
        _log(f"[PROC] 기존 주문 조회: {len(order_map)}건")

        # ─── 5단계: 분류 ───
        reupload_rows = []
        import_rows   = []
        new_rows      = []

        for row in raw_rows:
            ono = row["existing_order_no"]
            if ono:
                (reupload_rows if ono in order_map else import_rows).append(row)
            else:
                new_rows.append(row)

        _log(f"[PROC] 분류: 재업로드={len(reupload_rows)}, 초기임포트={len(import_rows)}, 신규={len(new_rows)}")

        inserted = 0
        updated  = 0
        errors   = []
        activity_batch = []

        # ══════════════════════════════════════
        # A: 재업로드 (DB에 존재)
        # ══════════════════════════════════════
        for row in reupload_rows:
            try:
                order_no  = row["existing_order_no"]
                order_id  = order_map[order_no]["id"]
                ex_item   = next(
                    (it for it in item_map.get(order_id,[])
                     if it["product_name"] == row["product_name"]), None)

                if ex_item:
                    changes, patch = [], {}
                    ns = row["status"]
                    if ex_item["status"] != ns:
                        changes.append(f"상태:{ex_item['status']}→{ns}")
                        patch["status"] = ns
                        patch["status_history"] = (
                            (ex_item.get("status_history") or ex_item["status"]) + "→" + ns)
                    if ex_item["quantity"] != row["quantity"]:
                        changes.append(f"수량:{ex_item['quantity']}→{row['quantity']}")
                        patch["quantity"] = row["quantity"]
                    if row["color"] and ex_item.get("color") != row["color"]:
                        changes.append(f"색상:{ex_item.get('color','없음')}→{row['color']}")
                        patch["color"] = row["color"]

                    note = (f"[{now_str}] 재업로드: {', '.join(changes)}"
                            if changes else f"[{now_str}] 재업로드 (변경없음)")
                    patch["change_log"] = ((ex_item.get("change_log") or "") + "\n" + note).strip()

                    if patch:
                        supabase.table("order_items").update(patch).eq("id", ex_item["id"]).execute()
                        ex_item.update(patch)

                    activity_batch.append({
                        "event_type": "re_upload" if changes else "re_upload_no_change",
                        "order_no": order_no, "product_name": row["product_name"],
                        "manager_code": row["manager_code"],
                        "old_value": ex_item["status"], "new_value": ns,
                        "note": note, "upload_history_id": upload_id,
                    })
                    updated += 1
                else:
                    supabase.table("order_items").insert({
                        "order_id": order_id, "product_name": row["product_name"],
                        "quantity": row["quantity"], "color": row["color"] or None,
                        "status": row["status"], "barcode": row["barcode"] or None,
                        "brand": row["brand"] or None, "size": row["size"] or None,
                        "options": row["options"] or None,
                        "wholesale_price": row["wholesale"] or None,
                        "supplier": row["supplier"] or None,
                        "item_notes": row["item_notes"] or None,
                        "recipient_name": row["recipient_name"] or None,
                        "phone": row["phone"] or None, "address": row["address"] or None,
                        "buyer_user_id": row["bx_user_id"] or None,
                        "delivery_msg": row["delivery_msg"] or None,
                        "item_code": row["item_code"] or None,
                        "status_history": row["status"],
                        "change_log": f"[{now_str}] 기존 주문에 상품 추가",
                    }).execute()
                    activity_batch.append({
                        "event_type":"new_upload","order_no":order_no,
                        "product_name":row["product_name"],"manager_code":row["manager_code"],
                        "old_value":None,"new_value":row["status"],
                        "note":f"[{now_str}] 기존 주문에 상품 추가",
                        "upload_history_id":upload_id,
                    })
                    inserted += 1
            except Exception as e:
                errors.append(f"행 {row['row_idx']}: {e}")

        # ══════════════════════════════════════
        # B: 초기 임포트 — bulk INSERT
        # ══════════════════════════════════════
        if import_rows:
            _log(f"[PROC] 초기 임포트 {len(import_rows)}행 처리 중...")
            import_groups = defaultdict(list)
            for row in import_rows:
                import_groups[row["existing_order_no"]].append(row)

            orders_to_insert = []
            for order_no, items in import_groups.items():
                first = items[0]
                mgr_id = mgr_cache.get(first["manager_code"])
                buy_id = _get_buyer_id(first, buyer_cache)
                con_id = con_cache.get(first["consignor_name"]) if first["consignor_name"] else None
                if not mgr_id or not buy_id:
                    for it in items:
                        errors.append(f"행 {it['row_idx']}: manager/buyer 조회 실패")
                    continue
                orders_to_insert.append({
                    "order_no": order_no, "manager_id": mgr_id,
                    "buyer_id": buy_id, "consignor_id": con_id,
                    "order_date": first["order_date"], "status": first["status"],
                    "upload_history_id": upload_id,
                })

            for chunk in _chunks(orders_to_insert, BATCH):
                try:
                    supabase.table("orders").upsert(
                        chunk, on_conflict="order_no", ignore_duplicates=True
                    ).execute()
                except Exception:
                    for o in chunk:
                        try: supabase.table("orders").insert(o).execute()
                        except: pass

            new_ono_list = list(import_groups.keys())
            new_order_map = _bulk_fetch_orders(supabase, new_ono_list)

            items_to_insert = []
            act_to_add = []
            for order_no, items in import_groups.items():
                oi = new_order_map.get(order_no)
                if not oi:
                    for it in items:
                        errors.append(f"행 {it['row_idx']}: order_no={order_no} 생성 실패")
                    continue
                oid = oi["id"]
                for row in items:
                    items_to_insert.append({
                        "order_id": oid, "product_name": row["product_name"],
                        "quantity": row["quantity"], "color": row["color"] or None,
                        "status": row["status"], "barcode": row["barcode"] or None,
                        "brand": row["brand"] or None, "size": row["size"] or None,
                        "options": row["options"] or None,
                        "wholesale_price": row["wholesale"] or None,
                        "supplier": row["supplier"] or None,
                        "item_notes": row["item_notes"] or None,
                        "recipient_name": row["recipient_name"] or None,
                        "phone": row["phone"] or None, "address": row["address"] or None,
                        "buyer_user_id": row["bx_user_id"] or None,
                        "delivery_msg": row["delivery_msg"] or None,
                        "item_code": row["item_code"] or None,
                        "status_history": row["status"],
                        "change_log": f"[{now_str}] 초기 임포트",
                    })
                    act_to_add.append({
                        "event_type":"new_upload","order_no":order_no,
                        "product_name":row["product_name"],"manager_code":row["manager_code"],
                        "old_value":None,"new_value":row["status"],
                        "note":f"[{now_str}] 초기 임포트","upload_history_id":upload_id,
                    })

            _bulk_insert(supabase, "order_items", items_to_insert)
            activity_batch.extend(act_to_add)
            inserted += len(items_to_insert)

        # ══════════════════════════════════════
        # C: 신규 (order_no 없음) — 로컬 번호 생성
        # ══════════════════════════════════════
        if new_rows:
            _log(f"[PROC] 신규 주문 {len(new_rows)}행 처리 중 (그룹={len(new_groups)}개)...")

            # 그룹별 엔티티 ID 수집
            group_data = {}
            for (bn, cn, od, mc), items in new_groups.items():
                first = items[0]
                buy_id = _get_buyer_id(first, buyer_cache)
                con_id = con_cache.get(cn) if cn else None
                mgr_id = mgr_cache.get(mc)
                if not buy_id or not mgr_id:
                    for it in items:
                        errors.append(f"행 {it['row_idx']}: buyer/manager 조회 실패 (buyer={bn}, mgr={mc})")
                    continue
                gk = f"{bn}||{cn}||{od}||{mc}"
                group_data[gk] = {
                    "buyer_id":   buy_id,
                    "consignor_id": con_id,
                    "manager_code": mc,
                    "manager_id": mgr_id,
                    "buyer_id_str": str(buy_id),
                    "con_id_str": str(con_id) if con_id else None,
                }

            # 로컬에서 base_number 할당 (RPC 없음)
            _log("[PROC] 주문번호 생성 중 (로컬)...")
            base_num_map = _build_order_no_engine(supabase, group_data)
            _log(f"[PROC] 주문번호 생성 완료: {len(base_num_map)}개 그룹")

            # 주문번호 생성 함수
            def make_order_no(mc, od, base_num, is_consignment, seq_num, total_count):
                prefix = "#" if is_consignment else ""
                # od는 "2026-04-14" → "260414"
                try:
                    ymd = datetime.strptime(od, "%Y-%m-%d").strftime("%y%m%d")
                except Exception:
                    ymd = od.replace("-", "")[2:8]
                return f"{prefix}{mc}{ymd}-{base_num}({seq_num}/{total_count})"

            # 주문 및 아이템 배열 구성
            orders_to_insert = []
            items_meta = []  # (order_no, row)

            for (bn, cn, od, mc), items in new_groups.items():
                gk = f"{bn}||{cn}||{od}||{mc}"
                if gk not in base_num_map:
                    for it in items:
                        errors.append(f"행 {it['row_idx']}: base_number 없음")
                    continue

                gdata = group_data.get(gk)
                if not gdata:
                    for it in items:
                        errors.append(f"행 {it['row_idx']}: 그룹 데이터 없음")
                    continue

                base_num = base_num_map[gk]
                total = len(items)

                for idx, row in enumerate(items, 1):
                    is_consignment = row["is_consignment"]
                    seq = row["seq_num"]
                    order_no = make_order_no(mc, od, base_num, is_consignment, seq, total)
                    row["_order_no"] = order_no

                    orders_to_insert.append({
                        "order_no":          order_no,
                        "manager_id":        str(gdata["manager_id"]),
                        "buyer_id":          str(gdata["buyer_id"]),
                        "consignor_id":      gdata["con_id_str"],
                        "order_date":        od,
                        "status":            row["status"],
                        "upload_history_id": upload_id,
                    })
                    items_meta.append((order_no, row))

            _log(f"[PROC] 주문 INSERT 중 ({len(orders_to_insert)}건)...")
            # Bulk insert orders (ignore duplicates)
            for chunk in _chunks(orders_to_insert, BATCH):
                try:
                    supabase.table("orders").upsert(
                        chunk, on_conflict="order_no", ignore_duplicates=True
                    ).execute()
                except Exception as e:
                    _log(f"[WARN] orders upsert error: {e}")
                    for o in chunk:
                        try: supabase.table("orders").insert(o).execute()
                        except: pass

            # Fetch inserted order IDs
            all_new_nos = [o["order_no"] for o in orders_to_insert]
            new_order_id_map = {}
            for chunk in _chunks(all_new_nos, BATCH):
                data = supabase.table("orders").select("id,order_no").in_(
                    "order_no", chunk).execute().data or []
                for r in data: new_order_id_map[r["order_no"]] = r["id"]

            _log(f"[PROC] 아이템 INSERT 중 ({len(items_meta)}건)...")
            items_to_insert = []
            act_to_add = []
            for (order_no, row) in items_meta:
                oid = new_order_id_map.get(order_no)
                if not oid:
                    errors.append(f"행 {row['row_idx']}: order_id 조회 실패 {order_no}")
                    continue
                items_to_insert.append({
                    "order_id":        oid,
                    "product_name":    row["product_name"],
                    "quantity":        row["quantity"],
                    "color":           row["color"] or None,
                    "status":          row["status"],
                    "barcode":         row["barcode"] or None,
                    "brand":           row["brand"] or None,
                    "size":            row["size"] or None,
                    "options":         row["options"] or None,
                    "wholesale_price": row["wholesale"] or None,
                    "supplier":        row["supplier"] or None,
                    "item_notes":      row["item_notes"] or None,
                    "recipient_name":  row["recipient_name"] or None,
                    "phone":           row["phone"] or None,
                    "address":         row["address"] or None,
                    "buyer_user_id":   row["bx_user_id"] or None,
                    "delivery_msg":    row["delivery_msg"] or None,
                    "item_code":       row["item_code"] or None,
                    "status_history":  row["status"],
                    "change_log":      f"[{now_str}] 신규 등록 | 번호: {order_no}",
                })
                act_to_add.append({
                    "event_type":   "new_upload",
                    "order_no":     order_no,
                    "product_name": row["product_name"],
                    "manager_code": row["manager_code"],
                    "old_value":    None,
                    "new_value":    row["status"],
                    "note":         f"[{now_str}] 신규 등록",
                    "upload_history_id": upload_id,
                })

            _bulk_insert(supabase, "order_items", items_to_insert)
            activity_batch.extend(act_to_add)
            inserted += len(items_to_insert)
            _log(f"[PROC] 신규 처리 완료: orders={len(orders_to_insert)}, items={len(items_to_insert)}")

        # activity_log 대량 INSERT
        if activity_batch:
            _log(f"[PROC] activity_log INSERT 중 ({len(activity_batch)}건)...")
            _bulk_insert(supabase, "activity_log", activity_batch)

        supabase.table("upload_history").update({
            "status": "완료" if not errors else "완료(오류있음)",
            "rows_processed": len(raw_rows),
            "rows_inserted": inserted, "rows_updated": updated,
            "error_message": "\n".join(errors[:20]) if errors else None,
        }).eq("id", upload_id).execute()

        _log(f"[PROC] 완료: inserted={inserted}, updated={updated}, errors={len(errors)}")
        return {"success":True,"upload_id":upload_id,
                "inserted":inserted,"updated":updated,"errors":errors[:20]}

    except Exception as e:
        import traceback
        _log(f"[PROCESS ERROR] {e}\n{traceback.format_exc()}")
        if upload_id:
            try:
                supabase.table("upload_history").update({
                    "status":"실패","error_message":str(e)
                }).eq("id", upload_id).execute()
            except: pass
        return {"success":False,"error":str(e),"upload_id":upload_id}


# ──────────────────────────────────────
# 엑셀 내보내기
# ──────────────────────────────────────
def export_to_excel(rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "주문목록"
    ws.append(["알파벳","미등록주문","주문일","아이디(주문)","고유번호",
               "주문자명","위탁자명","브랜드","상품명","색상","사이즈","수량",
               "상가","도매가","미송","비고","이름","전화번호","주소",
               "아이디(구매)","배송메세지","코드","상품상태"])
    fills = {s: PatternFill(start_color=c,end_color=c,fill_type="solid")
             for s,c in STATUS_COLORS.items()}
    for row in rows:
        ws.append([
            row.get("manager_code",""), row.get("barcode",""), row.get("order_date",""),
            row.get("buyer_user_id_ref",""), row.get("order_no",""),
            row.get("buyer_name",""), row.get("consignor_name",""), row.get("brand",""),
            row.get("product_name",""), row.get("color",""), row.get("size",""),
            row.get("quantity",""), row.get("options",""), row.get("wholesale_price",""),
            row.get("supplier",""), row.get("item_notes",""), row.get("recipient_name",""),
            row.get("phone",""), row.get("address",""), row.get("buyer_user_id",""),
            row.get("delivery_msg",""), row.get("item_code",""), row.get("item_status",""),
        ])
        st = row.get("item_status","")
        if st in fills: ws.cell(ws.max_row, 9).fill = fills[st]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
